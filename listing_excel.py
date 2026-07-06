from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

EXCEL_PATH = Path(__file__).with_name('listings.xlsx')

HEADERS = [
    'Property Name',
    'Area',
    'Property Type',
    'Size (sqft)',
    'Rental Price',
    'Bedrooms',
    'Bathrooms',
    'Parking',
    'Price sorting',
    'Message Link',
]

LINK_FONT = Font(color='0563C1', underline='single')


def reset_listings_file(path=EXCEL_PATH):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(HEADERS)
    workbook.save(path)


def append_listing_row(fields, message_link, path=EXCEL_PATH):
    if path.exists():
        workbook = load_workbook(path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(HEADERS)

    sheet.append([
        fields.get('property_name'),
        fields.get('area'),
        fields.get('property_type'),
        fields.get('size_sqft'),
        fields.get('rental_price'),
        fields.get('bedrooms'),
        fields.get('bathrooms'),
        fields.get('parking'),
        fields.get('sorting_price'),
        message_link
    ])

    if message_link:
        link_cell = sheet.cell(row=sheet.max_row, column=len(HEADERS))
        link_cell.hyperlink = message_link
        link_cell.font = LINK_FONT

    workbook.save(path)
